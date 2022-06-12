from flask import Flask, render_template, request, redirect, url_for

import openpyxl

book = openpyxl.open("успеваемость.xlsx", read_only=True, data_only=True)
list_sheet_names = book.sheetnames

app = Flask(__name__)


@app.route('/', methods=['POST', 'GET'])
def index():
    return render_template("index.html", option=list_sheet_names)


@app.route('/table', methods=['POST', 'GET'])
def table():
    if request.method == 'POST':
        list_name = request.form.get("search_list")

        sheet = book.get_sheet_by_name(list_name)
        sheet_list = []
        for row in range(1, sheet.max_row + 1):
            row_list = []
            for col in range(0, sheet.max_column):
                row_list.append(sheet[row][col].value)
            sheet_list.append(row_list)

        html = '<table>'
        for row in sheet_list:
            html += '<tr>'
            for value in row:
                html += '<td>{}</td>'.format(value)
            html += '</tr>'
        html += '</table>'

        return html


if __name__ == "__main__":
    app.run(debug=True)
